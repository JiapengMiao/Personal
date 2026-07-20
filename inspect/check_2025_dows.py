import openpyxl
from datetime import datetime
wb = openpyxl.load_workbook(r"C:\Users\56558\Nutstore\1\我的坚果云\agent\Project-002-白银数据网页可视化\data\010源数据\白银所有数据.xlsx", read_only=True, data_only=True)
ws = wb["白银数据"]
# Find rows around 2025-01-01
from collections import Counter
dows_2025 = Counter()
count_2025 = 0
for r in range(3, ws.max_row+1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, datetime) and v.year == 2025:
        dows_2025[v.weekday()] += 1
        count_2025 += 1
    elif isinstance(v, datetime) and v.year > 2025:
        break
print(f"2025 rows: {count_2025}")
print(f"2025 weekday dist: {dict(sorted(dows_2025.items()))}")
# Also check 2026
dows_2026 = Counter()
count_2026 = 0
for r in range(3, ws.max_row+1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, datetime) and v.year == 2026:
        dows_2026[v.weekday()] += 1
        count_2026 += 1
print(f"2026 rows: {count_2026}")
print(f"2026 weekday dist: {dict(sorted(dows_2026.items()))}")
wb.close()
