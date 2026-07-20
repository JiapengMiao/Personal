import openpyxl
from datetime import datetime
wb = openpyxl.load_workbook(r"C:\Users\56558\Nutstore\1\我的坚果云\agent\Project-002-白银数据网页可视化\data\010源数据\白银所有数据.xlsx", read_only=True, data_only=True)
ws = wb["白银数据"]
# row 3 = first data row, last row = last data row
first_date = ws.cell(row=3, column=1).value
last_date = ws.cell(row=ws.max_row, column=1).value
print(f"First date: {first_date}")
print(f"Last date: {last_date}")
print(f"Total rows (data): {ws.max_row - 2}")
# Check a few rows to see if it's calendar days or trading days
from collections import Counter
dows = Counter()
for r in range(3, min(103, ws.max_row+1)):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, datetime):
        dows[v.weekday()] += 1
print(f"First 100 rows weekday distribution (0=Mon..6=Sun): {dict(sorted(dows.items()))}")
wb.close()
