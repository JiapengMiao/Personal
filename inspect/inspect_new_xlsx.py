import openpyxl, json, sys
wb = openpyxl.load_workbook(r"C:\Users\56558\Nutstore\1\我的坚果云\agent\Project-002-白银数据网页可视化\data\010源数据\白银所有数据.xlsx", read_only=True, data_only=True)
info = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    nrows = ws.max_row
    ncols = ws.max_column
    info[sn] = {"rows": nrows, "cols": ncols, "header": rows[0] if rows else None, "row2": rows[1] if len(rows)>1 else None, "row3": rows[2] if len(rows)>2 else None}
wb.close()
for sn, v in info.items():
    print(f"\n=== {sn} === rows={v['rows']} cols={v['cols']}")
    print(f"  header: {v['header']}")
    if v['row2']: print(f"  row2:   {v['row2']}")
    if v['row3']: print(f"  row3:   {v['row3']}")
