# -*- coding: utf-8 -*-
"""从白银每日报价 Excel 提取全部历史日期报价，输出 spot_quotes.json。"""
import json
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(r"C:\Users\56558\Nutstore\1\金属投研小组\MJP-苗嘉鹏\数据计算\白银报价\白银每日报价.xlsx")
OUT = Path(__file__).resolve().parent.parent / "web" / "public" / "data" / "spot_quotes.json"

# 报价方分类（按 Excel 行顺序）
SMELTER = {  # 上游冶炼厂 1-14
    "山东招金", "深圳江铜", "江铜铅锌", "山东黄金", "山金实业",
    "中铜国际", "济金国际", "豫光金铅", "琨邦（恒邦）", "金川",
    "上海盛鸿（银泰）", "江西和丰环保（富冶）", "铜陵有色", "贵研黄金",
}
TRADER = {  # 贸易商 15-42
    "上海全银", "上海王银", "上海云羽", "宁波浩顺", "宁波凯通",
    "旷银", "远大生水", "上海紫薇智汇", "上海启梧/思烜", "郴州雄风",
    "维科嘉丰", "银峰", "三水", "五锐", "上海九石", "海亮",
    "厦门建发", "靖升", "中博世金", "翔佰盛屯", "五矿产融",
    "江铜国贸", "海南控股", "紫金", "四川资源", "托克", "埃克森", "先导科技",
}
FUTURES_SUB = {  # 期货公司风险子 43-53
    "华泰长城", "国贸启润", "国信金阳", "国泰君安", "国联汇富",
    "东证润和", "银河德睿", "兴证", "中信证券", "金瑞前海", "中信寰球",
}
BANK = {  # 银行 54-55
    "中信银行", "浙商银行（宁波）", "浙商银行",
}


def classify(name: str) -> str:
    if name in SMELTER:
        return "smelter"
    if name in TRADER:
        return "trader"
    if name in FUTURES_SUB:
        return "futures"
    if name in BANK:
        return "bank"
    return "other"


def main():
    wb = load_workbook(str(XLSX), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    wb.close()

    # 按日期分块
    all_days: list[dict] = []
    current_date = None
    current_block: list = []

    for r in rows:
        if r[0] is not None:
            # 保存上一个块
            if current_date and current_block:
                all_days.append(_build_day(current_date, current_block))
            current_date = r[0]
            current_block = [r]
        else:
            current_block.append(r)
    # 最后一块
    if current_date and current_block:
        all_days.append(_build_day(current_date, current_block))

    # 按日期升序
    all_days.sort(key=lambda d: d["date"])

    result = {
        "dates": [d["date"] for d in all_days],
        "days": all_days,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(result, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"[OK] spot_quotes.json ({len(all_days)} 天)")


def _build_day(date_val, block: list) -> dict:
    date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)[:10]
    td_spread = block[0][9] if len(block[0]) > 9 else None

    quotes = []
    for r in block:
        name = r[1]
        if not name:
            continue
        quotes.append({
            "name": name,
            "cat": classify(name),
            "fSh": str(r[2]).strip() if r[2] else None,
            "fGd": str(r[3]).strip() if r[3] else None,
            "fCt": str(r[4]).strip() if r[4] else None,
            "sSh": str(r[5]).strip() if r[5] else None,
            "sGd": str(r[6]).strip() if r[6] else None,
            "sCt": str(r[7]).strip() if r[7] else None,
            "note": str(r[8]).strip() if r[8] else None,
        })
    quoted = sum(1 for q in quotes if any([q["fSh"], q["fGd"], q["fCt"], q["sSh"], q["sGd"], q["sCt"]]))
    return {
        "date": date_str,
        "tdSpread": str(td_spread).strip() if td_spread else None,
        "count": len(quotes),
        "quoted": quoted,
        "quotes": quotes,
    }


if __name__ == "__main__":
    main()
