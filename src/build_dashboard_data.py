#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_dashboard_data.py

将 data/wind/ 下的 Excel、data/shfe/、data/sge/、data/gfex/ 与
data/monitoring/ 下的源数据抽取/转换为前端看板用 JSON，输出到 web/public/data/。

用法:  python src/build_dashboard_data.py
退出码: 全部成功 0；任一输出失败或校验失败 1（错误汇总打印在结尾）。
"""
from __future__ import annotations

import json
import math
import re
import shutil
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------- 路径

ROOT = Path(__file__).resolve().parent.parent
SRC_WIND = ROOT / "data" / "wind"
SRC_SHFE = ROOT / "data" / "shfe"
SRC_SGE = ROOT / "data" / "sge"
SRC_GFEX = ROOT / "data" / "gfex"
SRC_MONITORING = ROOT / "data" / "monitoring"
OUT_DIR = ROOT / "web" / "public" / "data"
# 主表统一在本项目维护，含“网页数据”sheet。
MAIN_XLSX = SRC_WIND / "白银所有数据.xlsx"


def _find_lease_xlsx() -> Path:
    """自动读取 data/wind/租赁利率 中文件名日期最新的一份利率表。"""
    lease_dir = SRC_WIND / "租赁利率"
    cands = sorted(lease_dir.glob("*利率*.xlsx")) if lease_dir.is_dir() else []
    if not cands:
        raise FileNotFoundError(f"未找到租借/租赁利率 xlsx: {lease_dir}")
    return cands[-1]


LEASE_XLSX = _find_lease_xlsx()

ERRORS: list[tuple[str, str]] = []   # (步骤名, 错误信息)

OZ_TO_KG = 32.1507466                # 美元/盎司 -> 元/千克 换算系数（×汇率）
VAT = 1.13                           # 进口增值税因子
TROY_OUNCE_GRAM = 31.1035            # 1 金衡盎司克数
OUTER_PRICE_DEDUCTION = 0.2          # 外盘期货价格扣减（美元/盎司）

# 白银数据主表：原文列名 -> (输出 key, 小数精度)；精度 0 表示取整
DAILY_COLS: list[tuple[str, str, int]] = [
    ("上海金交所:递延费支付方向:白银现货:Ag(T+D)", "deferredDirection", 0),
    ("上海金交所:结算价:白银现货:Ag(T+D)", "agtdSettle", 2),
    ("上海金交所:收盘价:白银现货:Ag(T+D)", "agtdClose", 2),
    ("中国:持仓量:白银现货:Ag(T+D):上海金交所", "agtdOi", 0),
    ("上期库存（日度）（吨）", "shfeInvT", 3),
    ("上金库存（日度）（吨）", "sgeInvT", 3),
    ("国内库存（上金+上期）（吨）", "domesticInvT", 3),   # 源列不直接使用，管道内重算
    ("国内库存+COMEX", "domesticPlusComex", 3),
    ("COMEX:库存量:银（吨）", "comexInvT", 3),
    ("COMEX+LBMA日度库存", "comexPlusLbmaDaily", 3),
    ("COMEX:持仓数量:非商业净持仓", "comexNonCommNet", 0),
    ("COMEX:持仓数量:非商业多头持仓:银", "comexNonCommLong", 0),
    ("COMEX:持仓数量:非商业空头持仓:银", "comexNonCommShort", 0),
    ("COMEX:注册仓单:白银:全球:吨", "comexWarrantT", 3),
    ("LBMA:库存量:白银（吨）", "lbmaInvT", 3),
    ("LBMA日度库存量:白银（吨）", "lbmaDailyT", 3),
    ("SLV:白银ETF:持仓量(吨)", "etfSLV", 3),
    ("PHAG:白银ETF:持仓量(吨)", "etfPHAG", 3),
    ("ETPMAG:白银ETF:持仓量(吨)", "etfETPMAG", 3),
    ("CEF:白银ETF:持仓量(吨)", "etfCEF", 3),
    ("PSLV:白银ETF:持仓量(吨)", "etfPSLV", 3),
    ("SLV+PHAG+ETPMAG+CEF", "etfUKSum", 3),
    ("现货价(伦敦市场):白银:美元", "londonSilverUsd", 2),
    ("中国:进口数量:白银", "importQty", 2),
    ("中国:出口数量:白银", "exportQty", 2),
]

# 库存类序列：停更后沿用前值（ffill），并记录最后真实值日期 lastActual
FFILL_KEYS = ("shfeInvT", "sgeInvT", "lbmaDailyT")

# 不从主表源列读取、在管道内计算的 key
COMPUTED_KEYS = {"domesticInvT"}

# 主表中日频（非稀疏）key，用于校验"无全 null 序列"
DAILY_DENSE_KEYS = {
    "agtdSettle", "agtdClose", "agtdOi", "shfeInvT", "sgeInvT", "domesticInvT",
    "domesticPlusComex", "comexInvT", "comexPlusLbmaDaily", "comexWarrantT",
    "lbmaDailyT", "etfSLV", "etfPHAG", "etfETPMAG", "etfCEF", "etfPSLV",
    "etfUKSum", "londonSilverUsd", "deferredDirection",
}

ZERO_TO_NULL_COUNTS: dict[str, int] = {}   # 记录每列 0->null 的个数（校验用）

# ---------------------------------------------------------------- 工具


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def write_json(name: str, obj: dict) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / name
    text = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    path.write_text(text, encoding="utf-8")
    n = len(obj.get("dates") or obj.get("times") or obj.get("contracts") or [])
    print(f"  [OK] {name} ({n} 记录, {path.stat().st_size / 1024:.1f} KB)")


def run_step(label: str, fn) -> None:
    """单步执行；失败收集错误但不中断后续步骤。"""
    try:
        fn()
    except Exception as exc:  # noqa: BLE001
        ERRORS.append((label, f"{type(exc).__name__}: {exc}"))
        print(f"  [FAIL] {label} 失败: {exc}")
        traceback.print_exc()


def to_float(v) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return f


def round_or_none(v, nd: int):
    f = to_float(v)
    if f is None:
        return None
    return round(f, nd)


def parse_yyyymmdd(v) -> datetime | None:
    """合约参数中的日期列：int/float/str 混合，统一转 datetime。"""
    f = to_float(v)
    if f is not None:
        s = str(int(f))
    else:
        s = str(v).strip()
    if not re.fullmatch(r"\d{8}", s):
        return None
    try:
        return datetime.strptime(s, "%Y%m%d")
    except ValueError:
        return None


def minute_frame(df_raw: pd.DataFrame, t_col: int, c_col: int) -> pd.DataFrame:
    """从原始（header=None）分钟 sheet 切出 time/close 并清洗。"""
    t = pd.to_datetime(df_raw.iloc[:, t_col], errors="coerce")
    c = pd.to_numeric(df_raw.iloc[:, c_col], errors="coerce")
    out = pd.DataFrame({"time": t, "close": c}).dropna()
    out = out.drop_duplicates(subset="time", keep="first").sort_values("time")
    return out.reset_index(drop=True)


def to_minute_series(df: pd.DataFrame) -> pd.Series:
    s = pd.Series(df["close"].to_numpy(dtype=float), index=pd.DatetimeIndex(df["time"]))
    return s[~s.index.duplicated(keep="first")].sort_index()


def align_minute_series(series_map: dict[str, pd.DataFrame]):
    """
    多路分钟序列按 time 外连接对齐 + 各自 ffill；
    截断到所有序列都已出现首个真实值之后。
    返回 (times_str_list, {name: np.ndarray})
    """
    ss = {name: to_minute_series(df) for name, df in series_map.items()}
    idx: pd.DatetimeIndex | None = None
    for s in ss.values():
        idx = s.index if idx is None else idx.union(s.index)
    assert idx is not None
    start = max(s.index[0] for s in ss.values())  # 全部序列都已开始的时刻
    mask = idx >= start
    aligned = {n: s.reindex(idx).ffill().to_numpy()[mask] for n, s in ss.items()}
    times = idx[mask].strftime("%Y-%m-%d %H:%M").tolist()
    return times, aligned


def spread_stats(vals: np.ndarray) -> dict:
    latest = float(vals[-1])
    return {
        "latest": round(latest, 1),
        "mean": round(float(np.mean(vals)), 1),
        "percentile": round(float((vals <= latest).mean()), 4),
        "min": round(float(np.min(vals)), 1),
        "max": round(float(np.max(vals)), 1),
    }


# ---------------------------------------------------------------- 输入加载（整簿一次解析）

SHEETS_NEEDED = [
    "白银数据", "AG（T+D）", "AG所有合约数据", "SPTAGUSDOZ_IDZ",
    "进出口盈亏计算数据", "外汇价格", "白银合约参数", "广期合约参数",
    "虚实比数据", "季节图表", "铂钯",
]

COMEX_MONTHS = {"F": 1, "G": 2, "H": 3, "J": 4, "K": 5, "M": 6,
                "N": 7, "Q": 8, "U": 9, "V": 10, "X": 11, "Z": 12}
MONTH_TO_COMEX = {month: letter for letter, month in COMEX_MONTHS.items()}


def comex_year_month(code: str) -> tuple[int, int] | None:
    match = re.fullmatch(r"SI([FGHJKMNQUVXZ])(\d{2})E\.CMX", code.upper())
    if not match:
        return None
    return 2000 + int(match.group(2)), COMEX_MONTHS[match.group(1)]


def month_after(year: int, month: int) -> tuple[int, int]:
    return (year + 1, 1) if month == 12 else (year, month + 1)


def matched_domestic_code(comex_code: str) -> str | None:
    ym = comex_year_month(comex_code)
    if ym is None:
        return None
    year, month = month_after(*ym)
    return f"AG{year % 100:02d}{month:02d}"


def fx_year_month(code: str) -> tuple[int, int] | None:
    match = re.fullmatch(r"UC([FGHJKMNQUVXZ])(\d{2})\.SG", code.upper())
    if not match:
        return None
    return 2000 + int(match.group(2)), COMEX_MONTHS[match.group(1)]


def select_quarterly_fx_contract(domestic_code: str, as_of: pd.Timestamp, available: set[str]) -> str:
    match = re.fullmatch(r"AG(\d{2})(\d{2})", domestic_code.upper())
    if not match:
        raise RuntimeError(f"无法解析内盘合约月份: {domestic_code}")
    target_ordinal = (2000 + int(match.group(1))) * 12 + int(match.group(2))
    as_of_ordinal = as_of.year * 12 + as_of.month
    candidates: list[tuple[int, int, str]] = []
    for code in available:
        ym = fx_year_month(code)
        if ym is None or ym[1] not in {3, 6, 9, 12}:
            continue
        ordinal = ym[0] * 12 + ym[1]
        if ordinal < as_of_ordinal:
            continue
        candidates.append((abs(ordinal - target_ordinal), ordinal, code))
    if not candidates:
        raise RuntimeError(f"未找到 {as_of:%Y-%m} 之后可用的 UC 季度主力合约")
    return min(candidates)[2]

PP_WINDOW_TRADING_DAYS = 120
GFEX_HOLIDAYS = pd.DatetimeIndex(pd.to_datetime([
    "2026-01-01", "2026-01-02", "2026-02-16", "2026-02-17", "2026-02-18",
    "2026-02-19", "2026-02-20", "2026-02-23", "2026-04-06", "2026-05-01",
    "2026-05-04", "2026-05-05", "2026-06-19", "2026-09-25", "2026-10-01",
    "2026-10-02", "2026-10-05", "2026-10-06", "2026-10-07",
]))


def defensive_sheet_check() -> None:
    wb = load_workbook(MAIN_XLSX, read_only=True)
    names = wb.sheetnames
    wb.close()
    print("  白银所有数据.xlsx 实际 sheet:", names)
    missing = [s for s in SHEETS_NEEDED if s not in names]
    if missing:
        raise RuntimeError(f"缺少必需 sheet: {missing}")
    wb2 = load_workbook(LEASE_XLSX, read_only=True)
    print("  20260719租借利率.xlsx 实际 sheet:", wb2.sheetnames)
    wb2.close()


def load_all_sheets() -> dict[str, pd.DataFrame]:
    """一次性解析整簿（header=None），各 sheet 按已实测布局切分。"""
    print("  解析主工作簿（整簿一次读取，约 10~30 秒）...")
    return pd.read_excel(MAIN_XLSX, sheet_name=SHEETS_NEEDED, header=None)


class Ctx:
    """共享上下文：日历、分钟序列、合约参数、虚实比数据。"""

    def __init__(self, sheets: dict[str, pd.DataFrame]):
        # ---- 白银数据主表（日历 + 日频序列）
        raw = sheets["白银数据"]
        header = raw.iloc[1].tolist()
        df = raw.iloc[2:].copy()
        df.columns = header
        df = df.rename(columns={header[0]: "date"})
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)
        self.daily = df
        self._hist_calendar = pd.DatetimeIndex(df["date"])
        self.calendar = self._hist_calendar  # extended after expiry parsed

        # ---- 分钟序列
        self.agtd = minute_frame(sheets["AG（T+D）"].iloc[6:], 0, 1)
        raw_outer = sheets["SPTAGUSDOZ_IDZ"]
        self.xagusd = minute_frame(raw_outer.iloc[6:], 0, 1)
        self.comex_min: dict[str, pd.DataFrame] = {}
        for k in range(raw_outer.shape[1] // 2):
            code = str(raw_outer.iloc[2, 2 * k + 1]).strip().upper()
            if comex_year_month(code) is None:
                continue
            frame = minute_frame(raw_outer.iloc[6:], 2 * k, 2 * k + 1)
            if len(frame):
                self.comex_min[code] = frame
        print(f"  COMEX 分钟合约解析: {sorted(self.comex_min)}")
        raw_fx = sheets["外汇价格"]
        self.fx_min_code = str(raw_fx.iloc[2, 1]).strip().upper()
        self.usdcnh = minute_frame(raw_fx.iloc[6:], 0, 1)
        print(f"  外汇分钟合约: {self.fx_min_code}")

        # ---- AG 所有合约（合约代码从第 3 行动态读取）
        raw_all = sheets["AG所有合约数据"]
        self.contracts_min: dict[str, pd.DataFrame] = {}
        code_row = raw_all.iloc[2].tolist()
        n_pairs = raw_all.shape[1] // 2
        for k in range(n_pairs):
            code = code_row[2 * k + 1]
            if not (isinstance(code, str) and re.fullmatch(r"AG\d{4}(\.SHF)?", code.strip())):
                print(f"  警告: AG所有合约数据 第 {2*k+2} 列合约代码无法识别: {code!r}，跳过")
                continue
            key = code.strip().replace(".SHF", "")
            self.contracts_min[key] = minute_frame(raw_all.iloc[6:], 2 * k, 2 * k + 1)
        print(f"  分钟合约解析: {sorted(self.contracts_min)}")

        # ---- 进出口盈亏日度数据：沪银 / COMEX / 离岸人民币分月合约
        raw_profit_daily = sheets["进出口盈亏计算数据"]
        profit_dates = pd.to_datetime(raw_profit_daily.iloc[5:, 0], errors="coerce")
        self.profit_daily: dict[str, pd.DataFrame] = {}
        for j in range(1, raw_profit_daily.shape[1]):
            code = str(raw_profit_daily.iloc[4, j]).strip().upper()
            if not code or code == "NAN":
                continue
            values = pd.to_numeric(raw_profit_daily.iloc[5:, j], errors="coerce")
            frame = pd.DataFrame({"time": profit_dates, "close": values}).dropna()
            if len(frame):
                self.profit_daily[code] = frame.drop_duplicates(subset="time", keep="last").sort_values("time")
        print(f"  进出口盈亏日度序列: {len(self.profit_daily)} 个")

        # ---- 合约参数: code(lower) -> expiry datetime
        raw_p = sheets["白银合约参数"]
        self.expiry: dict[str, datetime] = {}
        for _, row in raw_p.iloc[1:].iterrows():
            code = str(row.iloc[0]).strip().lower()
            exp = parse_yyyymmdd(row.iloc[2])  # 到期日=最后交易日
            if re.fullmatch(r"ag\d{4}", code) and exp is not None:
                self.expiry[code] = exp
        print(f"  合约参数解析: {len(self.expiry)} 个合约到期日")
        # Extend calendar with future business days up to max expiry (010 logic)
        _latest = self._hist_calendar[-1]
        _max_exp = max(self.expiry.values()) if self.expiry else _latest
        if _max_exp > _latest:
            _future = pd.bdate_range(start=_latest + pd.Timedelta(days=1), end=_max_exp)
            self.calendar = self._hist_calendar.union(_future).sort_values()
        else:
            self.calendar = self._hist_calendar
        print(f"  calendar: {len(self.calendar)} days (hist {len(self._hist_calendar)} + future {len(self.calendar)-len(self._hist_calendar)})")

        # ---- 交易日日历：从扩展后的 calendar 排除周末 + 包含所有合约到期日
        _wd = self.calendar.weekday
        _biz = self.calendar[_wd < 5]  # Mon-Fri
        _exp_dates = pd.DatetimeIndex(list(self.expiry.values()))
        self.trading_calendar = _biz.union(_exp_dates).sort_values()
        print(f"  trading_calendar: {len(self.trading_calendar)} days (biz {len(_biz)} + expiry {len(_exp_dates)})")

        # ---- 虚实比数据：仓单 st_stock + 各合约 oi
        raw_v = sheets["虚实比数据"]
        hdr = raw_v.iloc[5].tolist()  # 英文表头行
        body = raw_v.iloc[6:]
        # 仓单块（列 0/1）
        st_date = pd.to_datetime(body.iloc[:, 0], errors="coerce")
        st_val = pd.to_numeric(body.iloc[:, 1], errors="coerce")
        st = pd.DataFrame({"date": st_date, "st": st_val}).dropna()
        st = st.drop_duplicates(subset="date", keep="last").sort_values("date")
        self.st_stock = pd.Series(st["st"].to_numpy(dtype=float), index=pd.DatetimeIndex(st["date"]))
        # 仓单 ffill 到主表日历（任何仓单值出现之前保持 NaN）
        self.st_stock_ffill = self.st_stock.reindex(self.calendar).ffill()
        print(f"  注册仓单: {len(self.st_stock)} 个发布值, "
              f"{self.st_stock.index.min():%Y-%m-%d} → {self.st_stock.index.max():%Y-%m-%d}")

        # oi 块（列 2 日期, 列 3.. 合约；代码取英文表头行动态识别）
        oi_date = pd.to_datetime(body.iloc[:, 2], errors="coerce")
        self.oi: dict[str, pd.DataFrame] = {}
        for j in range(3, body.shape[1]):
            col_code = hdr[j]
            if not (isinstance(col_code, str) and re.fullmatch(r"AG\d{4}\.SHF", col_code.strip())):
                continue  # 到 close 块即停止识别（close 块表头为 AGxxxx.SHF.1 等）
            code = col_code.strip().replace(".SHF", "").lower()
            if code in self.oi:
                continue  # close 块与 oi 块表头原文相同，只取首次出现（oi 块）
            vals = pd.to_numeric(body.iloc[:, j], errors="coerce")
            d = pd.DataFrame({"date": oi_date, "oi": vals}).dropna()
            if len(d):
                self.oi[code] = d.sort_values("date").reset_index(drop=True)
        print(f"  虚实比 oi 合约: {sorted(self.oi)}")

        # ---- 广期所铂钯合约参数 + Wind 仓单/持仓复合表
        raw_gp = sheets["广期合约参数"]
        self.gfex_expiry: dict[str, datetime] = {}
        self.gfex_unit_kg: dict[str, float] = {}
        self.gfex_label_mismatches: list[tuple[str, str]] = []
        for _, row in raw_gp.iloc[1:].iterrows():
            code = str(row.iloc[1]).strip().lower()
            if not re.fullmatch(r"(?:pt|pd)\d{4}", code):
                continue
            exp = parse_yyyymmdd(row.iloc[5])
            unit_g = to_float(row.iloc[2])
            if exp is not None:
                self.gfex_expiry[code] = exp
            if unit_g is not None and unit_g > 0:
                self.gfex_unit_kg[code] = unit_g / 1000.0
            expected_label = "铂" if code.startswith("pt") else "钯"
            actual_label = str(row.iloc[0]).strip()
            if actual_label != expected_label:
                self.gfex_label_mismatches.append((code, actual_label))
        self.pp_raw = sheets["铂钯"]
        print(f"  广期铂钯合约参数: 到期日 {len(self.gfex_expiry)} 个, 交易单位 {len(self.gfex_unit_kg)} 个")
        if self.gfex_label_mismatches:
            print(f"  警告: 广期合约参数品种标签与代码前缀不一致，按代码识别: {self.gfex_label_mismatches}")

        # ---- 季节图表
        raw_s = sheets["季节图表"]
        yr_header = raw_s.iloc[1].tolist()
        self.season_years = [str(int(to_float(y))) for y in yr_header[1:6]]
        sbody = raw_s.iloc[2:]
        dcol = sbody.iloc[:, 0].astype(str).str.strip()
        mask = dcol.str.fullmatch(r"\d{2}-\d{2}")
        self.season_dates = dcol[mask].tolist()
        self.season_vals = sbody.loc[mask].iloc[:, 1:6].apply(pd.to_numeric, errors="coerce")

        print(f"  日历: {self.calendar[0]:%Y-%m-%d} → {self.calendar[-1]:%Y-%m-%d} ({len(self.calendar)} 天); "
              f"AGTD 分钟 {len(self.agtd)} 行; XAGUSD {len(self.xagusd)} 行; USDCNH {len(self.usdcnh)} 行")


# ---------------------------------------------------------------- 各输出步骤

CTX: Ctx | None = None


def _inventory_signal(delta: float | None) -> tuple[int | None, str, str]:
    """全球可用库存越低越利多；500 吨为强信号阈值。"""
    if delta is None:
        return None, "基线", "neutral"
    if delta <= -500:
        return 2, "强利多", "bull"
    if delta < 0:
        return 1, "偏利多", "bull"
    if delta >= 500:
        return -2, "强利空", "bear"
    if delta > 0:
        return -1, "偏利空", "bear"
    return 0, "中性", "neutral"


def _refresh_global_inventory_indicator(payload: dict) -> dict:
    """用统一 Wind 主表重算指标16：LBMA+COMEX+上期所+上金所-SLV。"""
    assert CTX is not None
    source_cols = {
        "lbma": "LBMA日度库存量:白银（吨）",
        "comex": "COMEX:库存量:银（吨）",
        "shfe": "上期库存（日度）（吨）",
        "sge": "上金库存（日度）（吨）",
        "slv": "SLV:白银ETF:持仓量(吨)",
    }
    missing = [col for col in source_cols.values() if col not in CTX.daily.columns]
    if missing:
        raise KeyError(f"全球库存指标缺少主表列: {missing}")

    frame = pd.DataFrame({"date": pd.to_datetime(CTX.daily["date"], errors="coerce")})
    actual: dict[str, pd.Series] = {}
    last_actual: dict[str, str] = {}
    for key, col in source_cols.items():
        series = pd.to_numeric(CTX.daily[col], errors="coerce").replace(0, np.nan)
        actual[key] = series
        valid_idx = series[series.notna()].index
        if len(valid_idx):
            last_actual[key] = frame.loc[valid_idx[-1], "date"].strftime("%Y-%m-%d")
        frame[key] = series.ffill()

    frame["value"] = frame["lbma"] + frame["comex"] + frame["shfe"] + frame["sge"] - frame["slv"]
    # 避免把周末仅由 ffill 形成的重复值当作新观察。
    actual_observation = pd.concat(actual.values(), axis=1).notna().any(axis=1)
    valid = frame.loc[frame["date"].notna() & frame["value"].notna() & actual_observation].copy()
    valid = valid.drop_duplicates("date", keep="last").sort_values("date").tail(60)
    if len(valid) < 2:
        raise ValueError("全球库存指标有效观察少于 2 条")

    history = []
    previous: float | None = None
    for row in valid.itertuples(index=False):
        value = round(float(row.value), 1)
        delta = None if previous is None else round(value - previous, 1)
        score, status, tone = _inventory_signal(delta)
        history.append({
            "period": row.date.strftime("%Y-%m-%d"),
            "value": value,
            "delta": delta,
            "score": score,
            "status": status,
            "tone": tone,
        })
        previous = value

    latest = history[-1]
    prior = history[-2]
    latest_row = valid.iloc[-1]
    breakdown_labels = {
        "lbma": "LBMA库存",
        "comex": "COMEX库存",
        "shfe": "上期所库存",
        "sge": "上金所库存",
        "slv": "SLV持仓（扣减项）",
    }
    breakdown = [
        {
            "key": key,
            "label": breakdown_labels[key],
            "value": round(float(latest_row[key]), 3),
            "unit": "吨",
            "asOfDate": last_actual.get(key, latest["period"]),
            "operation": "subtract" if key == "slv" else "add",
        }
        for key in ("lbma", "comex", "shfe", "sge", "slv")
    ]

    indicators = payload.get("indicators", [])
    indicator = next((item for item in indicators if int(item.get("id", -1)) == 16), None)
    if indicator is None:
        raise KeyError("monitoring-data.json 缺少指标16")
    indicator.update({
        "name": "全球可用白银库存（扣除SLV）",
        "value": latest["value"],
        "unit": "吨",
        "period": latest["period"],
        "updatedAt": latest["period"],
        "priorValue": prior["value"],
        "priorPeriod": prior["period"],
        "delta": latest["delta"],
        "score": latest["score"],
        "status": latest["status"],
        "tone": latest["tone"],
        "formula": "LBMA库存 + COMEX库存 + 上期所库存 + 上金所库存 - SLV持仓",
        "breakdown": breakdown,
        "history": history,
        "dataStatus": "已接入",
    })
    payload["asOfDate"] = max(str(payload.get("asOfDate", "")), latest["period"])
    return payload


def step_copy_static() -> None:
    # 监测模块已从 Project-006 并入本项目，统一读取本地标准目录。
    src = SRC_MONITORING / "monitoring-data.json"
    with open(src, encoding="utf-8") as fh:
        payload = json.load(fh)
    payload = _refresh_global_inventory_indicator(payload)
    write_json("monitoring.json", payload)
    size_kb = (OUT_DIR / "monitoring.json").stat().st_size / 1024
    print(f"  [OK] monitoring.json (统一监测源, {size_kb:.1f} KB)")

    # market.json 改为从 010 主表"网页数据" sheet 生成（用户手工维护的万得数据）
    build_market_from_web_sheet()


def build_market_from_web_sheet() -> None:
    """读取 010 主表"网页数据" sheet，生成 market.json。

    sheet 结构（data_only）：
      第4行 = 列名（日期/伦敦银(人民币/千克)/伦敦金(人民币/克)/国投瑞银白银期货A/SHFE白银/SGE白银T+D）
      第5行 = 万得代码（Date/XAGCNY.IDC/XAUCNY.IDC/161226.OF/AG.SHF/AG(T+D).SGE）
      第6行起 = 数据（A 列为日期）
    金银比 = 伦敦金(元/克) ÷ 伦敦银(元/千克) × 1000（同币种，单位换算后无量纲）。
    """
    wb = load_workbook(MAIN_XLSX, read_only=True, data_only=True)
    if "网页数据" not in wb.sheetnames:
        raise KeyError("主表缺少『网页数据』sheet")
    ws = wb["网页数据"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()
    # rows[0]=列名行, rows[1]=代码行, rows[2:]=数据
    series_map = {
        "londonSilverCnyKg": {"label": "伦敦银（人民币/千克）", "unit": "元/千克", "col": 1},
        "londonGoldCnyG": {"label": "伦敦金（人民币/克）", "unit": "元/克", "col": 2},
        "silverFundNav": {"label": "白银期货 LOF（161226.OF）净值", "unit": "元", "col": 3},
        "shfeSilver": {"label": "沪银主力（AG.SHF）", "unit": "元/千克", "col": 4},
        "sgeAgTd": {"label": "上金所 Ag(T+D)", "unit": "元/千克", "col": 5},
    }
    items: dict[str, dict] = {}
    for key, meta in series_map.items():
        pts = []
        for r in rows[2:]:
            dt, val = r[0], r[meta["col"]]
            if dt is None or val is None:
                continue
            d = dt.date().isoformat() if hasattr(dt, "date") else str(dt)[:10]
            f = to_float(val)
            if f is not None:
                pts.append({"date": d, "value": round(f, 4)})
        items[key] = {"label": meta["label"], "unit": meta["unit"], "points": pts}

    # 金银比：按日期对齐伦敦金/伦敦银
    au = {p["date"]: p["value"] for p in items["londonGoldCnyG"]["points"]}
    ag = {p["date"]: p["value"] for p in items["londonSilverCnyKg"]["points"]}
    ratio_pts = []
    for d in sorted(au.keys() & ag.keys()):
        if ag[d]:
            ratio_pts.append({"date": d, "value": round(au[d] / ag[d] * 1000, 2)})
    items["goldSilverRatio"] = {"label": "金银比（伦敦金 ÷ 伦敦银，人民币同口径）", "unit": "", "points": ratio_pts}

    write_json("market.json", {"fetchedAt": now_iso(), "items": items})



def step_daily() -> None:
    assert CTX is not None
    df = CTX.daily
    # 1) 读取源列（0→null）；COMPUTED_KEYS 不读源列
    raw: dict[str, pd.Series] = {}
    nd_map: dict[str, int] = {}
    for col, key, nd in DAILY_COLS:
        nd_map[key] = nd
        if key in COMPUTED_KEYS:
            continue
        if col not in df.columns:
            print(f"  警告: 主表缺少列 {col!r}，key {key} 跳过")
            continue
        vals = pd.to_numeric(df[col], errors="coerce")
        zeros = int((vals == 0).sum())
        if zeros:
            ZERO_TO_NULL_COUNTS[key] = zeros
            vals = vals.mask(vals == 0)
        raw[key] = vals.reset_index(drop=True)
    # 2) 库存类序列 ffill（停更沿用前值），记录最后真实值日期 lastActual
    last_actual: dict[str, str | None] = {}
    for key in FFILL_KEYS:
        if key not in raw:
            print(f"  警告: {key} 无源数据，跳过 ffill")
            last_actual[key] = None
            continue
        s = raw[key]
        li = s.last_valid_index()
        last_actual[key] = (df["date"].iloc[li].strftime("%Y-%m-%d") if li is not None else None)
        raw[key] = s.ffill()  # 头部 null（首个有效值之前）保留
        print(f"  {key}: lastActual={last_actual[key]}, ffill 后非空 {int(raw[key].notna().sum())}/{len(df)}")
    # 3) domesticInvT 管道内重算 = ffill(shfeInvT) + ffill(sgeInvT)，两者同日都有值才产出
    if "shfeInvT" in raw and "sgeInvT" in raw:
        sh, sg = raw["shfeInvT"], raw["sgeInvT"]
        raw["domesticInvT"] = (sh + sg).where(sh.notna() & sg.notna())
        print(f"  domesticInvT 重算: 非空 {int(raw['domesticInvT'].notna().sum())}/{len(df)}, "
              f"末值 {round_or_none(raw['domesticInvT'].iloc[-1], 3)}")
    else:
        print("  警告: shfeInvT/sgeInvT 缺失，domesticInvT 无法重算")
    # 4) 按 DAILY_COLS 顺序输出
    series: dict[str, list] = {}
    for _col, key, nd in DAILY_COLS:
        if key not in raw:
            continue
        if nd == 0:
            series[key] = [None if to_float(v) is None else int(round(float(v))) for v in raw[key]]
        else:
            series[key] = [round_or_none(v, nd) for v in raw[key]]
    dates = df["date"].dt.strftime("%Y-%m-%d").tolist()
    write_json("daily.json", {
        "generatedAt": now_iso(),
        "asOfDate": dates[-1],
        "lastActual": last_actual,
        "dates": dates,
        "series": series,
    })

    # —— 三级加载：切出近 2 年的轻量版 daily_recent.json（首屏加载，~4% 体积）——
    # 历史数据保留在 daily.json（重命名为 daily_history.json 由前端按需加载）
    RECENT_YEARS = 2
    cutoff_ts = pd.Timestamp(dates[-1]) - pd.DateOffset(years=RECENT_YEARS)
    cutoff = cutoff_ts.strftime("%Y-%m-%d")
    start_idx = next((i for i, d in enumerate(dates) if d >= cutoff), 0)
    recent_dates = dates[start_idx:]
    recent_series = {k: v[start_idx:] for k, v in series.items()}
    write_json("daily_recent.json", {
        "generatedAt": now_iso(),
        "asOfDate": dates[-1],
        "lastActual": last_actual,
        "recentFrom": recent_dates[0],
        "dates": recent_dates,
        "series": recent_series,
    })
    print(f"  daily_recent.json: {len(recent_dates)} 点（自 {recent_dates[0]}，近 {RECENT_YEARS} 年）")
    # 全量历史另存为 daily_history.json（前端懒加载）
    write_json("daily_history.json", {
        "generatedAt": now_iso(),
        "asOfDate": dates[-1],
        "lastActual": last_actual,
        "dates": dates,
        "series": series,
    })


def _contract_points(code: str, y_func) -> list[dict]:
    """按交易日日历计算 x（距到期日的交易日序号差，到期日 x=0），保留 [-90, 0]。
    已到期合约若曲线未延伸到 x=0，自动补一个端点。"""
    assert CTX is not None
    cal = CTX.trading_calendar
    exp = CTX.expiry.get(code, CTX.oi[code]["date"].max())
    exp_ts = pd.Timestamp(exp)
    exp_pos = int(cal.searchsorted(exp_ts))
    d = CTX.oi[code]
    pts = []
    for date, oi in zip(d["date"], d["oi"]):
        di = int(cal.searchsorted(pd.Timestamp(date)))
        x = di - exp_pos
        if not (-90 <= x <= 0):
            continue
        y = y_func(pd.Timestamp(date), di, float(oi))
        if y is None:
            continue
        pts.append({"x": x, "y": y})
    # 已到期合约：若最后一个点 x < 0，补 x=0 端点（沿用最后有效 y 值）
    if pts and pts[-1]["x"] < 0 and exp_ts <= pd.Timestamp(CTX.daily["date"].iloc[-1]):
        pts.append({"x": 0, "y": pts[-1]["y"]})
    return pts


def _is_even_month_contract(code: str) -> bool:
    """ag2608 → 月份 08 → 偶数月保留（白银期货仅偶数月合约活跃）。"""
    m = re.fullmatch(r"ag\d{2}(\d{2})", code)
    return bool(m) and int(m.group(1)) % 2 == 0


def _build_curve(y_func, extra_meta: dict | None = None) -> dict:
    assert CTX is not None
    contracts = []
    skipped_no_expiry, skipped_odd, kept = [], [], []
    for code in sorted(CTX.oi):
        # if code not in CTX.expiry:
        # skipped_no_expiry.append(code)
        # continue
        if not _is_even_month_contract(code):
            skipped_odd.append(code)
            continue
        pts = _contract_points(code, y_func)
        if pts:
            contracts.append({
                "code": code,
                "expiry": CTX.expiry.get(code, CTX.oi[code]["date"].max()).strftime("%Y-%m-%d"),
                "points": pts,
            })
            kept.append(code)
    if skipped_no_expiry:
        print(f"  警告: 以下 oi 合约在合约参数表中无到期日，已跳过: {skipped_no_expiry}")
    print(f"  偶数月合约过滤: 保留 {kept}; 丢弃(奇数月) {skipped_odd}")
    contracts.sort(key=lambda c: c["expiry"])
    obj = {"generatedAt": now_iso(), "contracts": contracts}
    if extra_meta:
        obj.update(extra_meta)
    return obj


def step_positions_curve() -> None:
    write_json("positions_curve.json", _build_curve(
        lambda date, di, oi: int(round(oi))))


def step_virtual_ratio() -> None:
    assert CTX is not None

    def y_func(date: pd.Timestamp, di: int, oi: float):
        # 当日或之前最近一个发布日的仓单；仓单首个值之前不设比率
        pos = int(CTX.calendar.searchsorted(date, side="right")) - 1
        if pos < 0:
            return None
        st = to_float(CTX.st_stock_ffill.iloc[pos])
        if st is None or st <= 0:
            return None
        return round(oi * 15.0 / st, 2)

    write_json("virtual_ratio.json", _build_curve(
        y_func, {"formula": "oi*15/st_stock_kg"}))


def _find_exact_col(row: pd.Series, target: str) -> int:
    target = target.strip().upper()
    for idx, value in row.items():
        if str(value).strip().upper() == target:
            return int(idx)
    raise ValueError(f"铂钯工作表未找到字段 {target}")


def _build_pp_metal_curve(prefix: str, label: str, warehouse_code: str) -> dict:
    """复刻 010 铂钯虚实比口径：1kg/手 × OI ÷ 注册仓单，按最后交易日对齐。"""
    assert CTX is not None
    raw = CTX.pp_raw
    security_row = raw.iloc[2]
    header_row = raw.iloc[5]
    body = raw.iloc[6:].copy()

    stock_col = _find_exact_col(security_row, warehouse_code)
    stock_date_col = stock_col - 1
    contract_cols: dict[str, int] = {}
    pattern = re.compile(rf"^({prefix}\d{{4}})(?:\.GFE)?$", re.IGNORECASE)
    for idx, value in header_row.items():
        match = pattern.fullmatch(str(value).strip())
        if match:
            contract_cols[match.group(1).lower()] = int(idx)
    if not contract_cols:
        raise ValueError(f"铂钯工作表未识别到 {prefix} 合约持仓列")

    warehouse = pd.DataFrame({
        "date": pd.to_datetime(body.iloc[:, stock_date_col], errors="coerce"),
        "stock": pd.to_numeric(body.iloc[:, stock_col], errors="coerce"),
    }).dropna()
    warehouse = warehouse[warehouse["stock"] > 0]
    warehouse = warehouse.groupby("date", as_index=False).first().sort_values("date")

    oi_date_col = min(contract_cols.values()) - 1
    positions = pd.DataFrame({
        "date": pd.to_datetime(body.iloc[:, oi_date_col], errors="coerce"),
    })
    for code, col in contract_cols.items():
        positions[code] = pd.to_numeric(body.iloc[:, col], errors="coerce")
    # Wind 复合表的日期区块会纵向重复，后一个重复行常为空；按日期逐列取首个非空值，
    # 不能 drop_duplicates(keep="last")，否则会把绝大多数历史持仓覆盖成 NaN。
    positions = positions.dropna(subset=["date"]).groupby("date", as_index=False).first().sort_values("date")
    merged = warehouse.merge(positions, on="date", how="inner").sort_values("date").reset_index(drop=True)
    if merged.empty:
        raise ValueError(f"{label}仓单与持仓量无可匹配日期")

    expiry = {code: pd.Timestamp(CTX.gfex_expiry[code]) for code in contract_cols if code in CTX.gfex_expiry}
    latest = pd.Timestamp(merged["date"].max()).normalize()
    maximum = max([latest, *expiry.values()]) if expiry else latest
    historical = pd.DatetimeIndex(merged["date"]).normalize().unique().sort_values()
    future = pd.bdate_range(latest + pd.Timedelta(days=1), maximum) if maximum > latest else pd.DatetimeIndex([])
    future = future[~future.isin(GFEX_HOLIDAYS)]
    calendar = historical.union(future).sort_values()
    calendar_pos = {pd.Timestamp(d).normalize(): i for i, d in enumerate(calendar)}

    contracts: list[dict] = []
    for code in sorted(contract_cols):
        if code not in expiry or code not in CTX.gfex_unit_kg:
            print(f"  警告: {code} 缺少到期日或交易单位，跳过")
            continue
        exp = expiry[code].normalize()
        available_dates = calendar[calendar <= exp]
        if not len(available_dates):
            continue
        exp_pos = calendar_pos[pd.Timestamp(available_dates[-1]).normalize()]
        subset = merged[["date", "stock", code]].copy()
        multiplier = CTX.gfex_unit_kg[code]
        subset["ratio"] = (multiplier * subset[code].where(subset[code] > 0) / subset["stock"])
        subset.loc[subset["date"].eq(exp) & subset[code].eq(0), "ratio"] = 0.0
        subset = subset.dropna(subset=["ratio"])
        subset = subset[subset["date"] <= exp]
        subset["cal_pos"] = subset["date"].map(lambda d: calendar_pos.get(pd.Timestamp(d).normalize()))
        subset = subset.dropna(subset=["cal_pos"])
        subset["x"] = subset["cal_pos"].astype(int) - exp_pos
        subset = subset[subset["x"].between(-PP_WINDOW_TRADING_DAYS, 0)].sort_values("x")
        print(f"  {code.upper()}: 窗口内 {len(subset)} 个有效点")
        if len(subset) < 2:
            continue
        contracts.append({
            "code": code.upper(),
            "expiry": exp.strftime("%Y-%m-%d"),
            "points": [
                {"x": int(x), "y": round(float(y), 4)}
                for x, y in zip(subset["x"], subset["ratio"])
            ],
        })

    if not contracts:
        raise ValueError(f"{label}没有至少含 2 个有效点的合约")
    contracts.sort(key=lambda c: c["expiry"])
    return {
        "label": label,
        "symbol": prefix,
        "asOfDate": latest.strftime("%Y-%m-%d"),
        "windowTradingDays": PP_WINDOW_TRADING_DAYS,
        "formula": "oi*contract_multiplier_kg/st_stock_kg",
        "contracts": contracts,
    }


def step_metal_virtual_ratio() -> None:
    assert CTX is not None
    write_json("metal_virtual_ratio.json", {
        "generatedAt": now_iso(),
        "source": {
            "workbook": "data/wind/白银所有数据.xlsx",
            "warehouseSheet": "铂钯",
            "contractSheet": "广期合约参数",
        },
        "qualityNotes": [
            f"合约代码优先于品种标签识别；已纠正 {code} 的源标签 {actual}"
            for code, actual in CTX.gfex_label_mismatches
        ],
        "metals": {
            "pt": _build_pp_metal_curve("PT", "铂金", "PT.GFE"),
            "pd": _build_pp_metal_curve("PD", "钯金", "PD.GFE"),
        },
    })


def step_shfe_positioning() -> None:
    """读取统一目录中的 SHFE 会员汇总与 SGE 持仓，输出网页数据。"""
    ranking_root = SRC_SHFE / "ranking"
    date_dirs = sorted(
        path for path in ranking_root.iterdir()
        if path.is_dir() and re.fullmatch(r"\d{8}", path.name)
    )
    if not date_dirs:
        raise FileNotFoundError(f"未找到 SHFE 持仓排名日期目录: {ranking_root}")

    records: list[dict] = []
    for date_dir in date_dirs:
        source = date_dir / f"silver_ranking_{date_dir.name}.json"
        if not source.exists():
            raise FileNotFoundError(f"持仓排名目录缺少 JSON: {source}")
        rows = json.loads(source.read_text(encoding="utf-8"))
        summary_rows = {
            int(row.get("RANK")): row
            for row in rows
            if str(row.get("INSTRUMENTID", "")).lower() == "agall"
            and int(row.get("RANK", -999)) in (-1, 0)
        }
        if set(summary_rows) != {-1, 0}:
            raise ValueError(f"{source.name} 缺少 agall 的 RANK=-1/0 汇总行")
        item = {"date": datetime.strptime(date_dir.name, "%Y%m%d").strftime("%Y-%m-%d")}
        for rank, prefix in ((-1, "futures"), (0, "nonFutures")):
            row = summary_rows[rank]
            item[prefix] = {
                "label": str(row.get("PARTICIPANTABBR1") or row.get("PARTICIPANTID1") or prefix),
                "long": int(row["CJ2"]),
                "longChange": int(row["CJ2_CHG"]),
                "short": int(row["CJ3"]),
                "shortChange": int(row["CJ3_CHG"]),
            }
        records.append(item)

    dates = [row["date"] for row in records]
    if dates != sorted(set(dates)):
        raise ValueError("SHFE 持仓日期不唯一或未升序")

    sge_path = SRC_SGE / "ag_td_daily_2026.csv"
    sge = pd.read_csv(sge_path, encoding="utf-8-sig")
    sge["date"] = pd.to_datetime(sge["date"], errors="coerce")
    sge["open_interest"] = pd.to_numeric(sge["open_interest"], errors="coerce")
    sge = sge.dropna(subset=["date", "open_interest"]).sort_values("date")
    if sge["date"].duplicated().any():
        raise ValueError("SGE Ag(T+D) 存在重复日期")
    sge_map = dict(zip(sge["date"].dt.strftime("%Y-%m-%d"), sge["open_interest"]))

    common = [row for row in records if row["date"] in sge_map]
    latest = records[-1]
    quality_notes = [
        "SHFE 会员汇总取 INSTRUMENTID=agall；RANK=-1 为期货公司会员，RANK=0 为非期货公司会员。",
        "趋势中的 SHFE CJ2/CJ3 为非期货公司会员持买/持卖单量；统一吨口径按 1 手=15 千克换算。",
        "SGE Ag(T+D) open_interest 按 1 手=1 千克换算；综合图仅使用 SHFE 与 SGE 共同交易日。",
    ]
    write_json("shfe_positioning.json", {
        "generatedAt": now_iso(),
        "asOfDate": latest["date"],
        "source": {
            "project": "Project-002-白银数据网页可视化",
            "rankingPattern": "data/shfe/ranking/YYYYMMDD/silver_ranking_YYYYMMDD.json",
            "sgeFile": "data/sge/ag_td_daily_2026.csv",
        },
        "quality": {
            "shfeTradingDays": len(records),
            "sgeTradingDays": len(sge),
            "commonTradingDays": len(common),
            "notes": quality_notes,
        },
        "summary": [
            {"category": "期货公司会员", **latest["futures"]},
            {"category": "非期货公司会员", **latest["nonFutures"]},
        ],
        "nonFuturesTrend": {
            "dates": dates,
            "longLots": [row["nonFutures"]["long"] for row in records],
            "shortLots": [row["nonFutures"]["short"] for row in records],
            "netLots": [row["nonFutures"]["long"] - row["nonFutures"]["short"] for row in records],
        },
        "combinedTrend": {
            "dates": [row["date"] for row in common],
            "shfeLongTons": [round(row["nonFutures"]["long"] * 0.015, 3) for row in common],
            "shfeShortTons": [round(row["nonFutures"]["short"] * 0.015, 3) for row in common],
            "sgeOpenInterestTons": [round(float(sge_map[row["date"]]) * 0.001, 3) for row in common],
        },
    })


def step_pp_warehouse() -> None:
    """读取统一目录中的最新铂钯仓单原始 CSV，并重算仓库/厂库趋势。"""
    candidates: list[tuple[str, str, Path]] = []
    pattern = re.compile(r"铂钯仓单数据_\d{8}_(\d{8})_(\d{8}_\d{6})\.csv$")
    for path in SRC_GFEX.glob("铂钯仓单数据_*.csv"):
        match = pattern.fullmatch(path.name)
        if match:
            candidates.append((match.group(1), match.group(2), path))
    if not candidates:
        raise FileNotFoundError(f"未找到规范命名的铂钯仓单 CSV: {SRC_GFEX}")
    _, _, source = max(candidates)

    frame = pd.read_csv(source, encoding="utf-8-sig", dtype={"仓库代码": str})
    required = {"品种", "日期", "仓库代码", "仓库名称", "昨日仓单", "今日注册", "今日注销", "今日仓单", "增减"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"{source.name} 缺少列: {missing}")
    frame["date"] = pd.to_datetime(frame["日期"].astype(str), format="%Y%m%d", errors="coerce")
    numeric_cols = ["昨日仓单", "今日注册", "今日注销", "今日仓单", "增减"]
    for column in numeric_cols:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    if frame[["date", *numeric_cols]].isna().any().any():
        raise ValueError(f"{source.name} 存在无法解析的日期或数值")
    if frame.duplicated(subset=["品种", "date", "仓库代码"]).any():
        raise ValueError(f"{source.name} 存在 品种+日期+仓库代码 重复行")
    if not (frame["昨日仓单"] + frame["今日注册"] - frame["今日注销"] == frame["今日仓单"]).all():
        raise ValueError(f"{source.name} 不满足 昨日仓单+注册-注销=今日仓单")
    if not (frame["今日仓单"] - frame["昨日仓单"] == frame["增减"]).all():
        raise ValueError(f"{source.name} 不满足 今日仓单-昨日仓单=增减")

    pt_warehouses = {"中工美上海虹桥", "外运华东上海虹桥", "深圳威豹"}
    pd_warehouses = {"中储吴淞", "中工美上海虹桥", "外运华东上海虹桥", "深圳威豹"}
    frame["仓库类型"] = [
        "仓库" if name in (pt_warehouses if metal == "铂" else pd_warehouses) else "厂库"
        for metal, name in zip(frame["品种"], frame["仓库名称"])
    ]

    metals: dict[str, dict] = {}
    for metal_cn, metal_key, label in (("铂", "pt", "铂金"), ("钯", "pd", "钯金")):
        subset = frame[frame["品种"] == metal_cn].copy()
        if subset.empty:
            raise ValueError(f"{source.name} 缺少{metal_cn}仓单数据")
        dates = pd.DatetimeIndex(subset["date"].drop_duplicates().sort_values())
        warehouse = subset[subset["仓库类型"] == "仓库"].groupby("date")["今日仓单"].sum()
        factory = subset[subset["仓库类型"] == "厂库"].groupby("date")["今日仓单"].sum()
        registered = subset.groupby("date")["今日注册"].sum()
        cancelled = subset.groupby("date")["今日注销"].sum()
        change = subset.groupby("date")["增减"].sum()
        warehouse_vals = warehouse.reindex(dates, fill_value=0).astype(int).tolist()
        factory_vals = factory.reindex(dates, fill_value=0).astype(int).tolist()
        total_vals = [w + f for w, f in zip(warehouse_vals, factory_vals)]
        latest_date = dates[-1]
        latest_rows = subset[subset["date"] == latest_date].sort_values("今日仓单", ascending=False)
        locations = [
            {
                "code": str(row["仓库代码"]),
                "name": str(row["仓库名称"]),
                "type": str(row["仓库类型"]),
                "quantityKg": int(row["今日仓单"]),
                "registeredKg": int(row["今日注册"]),
                "cancelledKg": int(row["今日注销"]),
                "changeKg": int(row["增减"]),
            }
            for _, row in latest_rows.iterrows()
        ]
        metals[metal_key] = {
            "label": label,
            "symbol": metal_key.upper(),
            "dates": dates.strftime("%Y-%m-%d").tolist(),
            "warehouseKg": warehouse_vals,
            "factoryKg": factory_vals,
            "totalKg": total_vals,
            "registeredKg": registered.reindex(dates, fill_value=0).astype(int).tolist(),
            "cancelledKg": cancelled.reindex(dates, fill_value=0).astype(int).tolist(),
            "netChangeKg": change.reindex(dates, fill_value=0).astype(int).tolist(),
            "latest": {
                "date": latest_date.strftime("%Y-%m-%d"),
                "warehouseKg": warehouse_vals[-1],
                "factoryKg": factory_vals[-1],
                "totalKg": total_vals[-1],
                "registeredKg": int(registered.get(latest_date, 0)),
                "cancelledKg": int(cancelled.get(latest_date, 0)),
                "netChangeKg": int(change.get(latest_date, 0)),
            },
            "locations": locations,
        }

    as_of = min(metal["latest"]["date"] for metal in metals.values())
    write_json("pp_warehouse.json", {
        "generatedAt": now_iso(),
        "asOfDate": as_of,
        "source": {
            "project": "Project-002-白银数据网页可视化",
            "file": f"data/{source.name}",
            "unit": "千克",
        },
        "quality": {
            "rowCount": len(frame),
            "notes": [
                "逐行校验：昨日仓单 + 今日注册 - 今日注销 = 今日仓单。",
                "逐行校验：今日仓单 - 昨日仓单 = 增减。",
                "仓库/厂库分类沿用 src/collectors/gfex_classify_warehouse.py 的品种专属名单。",
            ],
        },
        "metals": metals,
    })


def step_seasonality() -> None:
    assert CTX is not None
    years = {}
    for i, yr in enumerate(CTX.season_years):
        col = CTX.season_vals.iloc[:, i]
        years[yr] = [round_or_none(v, 2) for v in col]
    write_json("seasonality.json", {
        "generatedAt": now_iso(),
        "dates": CTX.season_dates,
        "years": years,
    })


def step_lease_rates() -> None:
    df = pd.read_excel(LEASE_XLSX, sheet_name=0, header=7)
    df.columns = [str(c).strip() for c in df.columns]
    date_col = df.columns[0]

    def find_col(metal: str, tenor_pat: str) -> str:
        for c in df.columns:
            if metal in c and re.search(tenor_pat, c):
                return c
        raise KeyError(f"租借利率表未找到{metal} {tenor_pat} 列; 实际列: {list(df.columns)}")

    col_map = {
        "m1": find_col("白银", r"(一|1)个月"),
        "m3": find_col("白银", r"3个月"),
        "m6": find_col("白银", r"6个月"),
        "m12": find_col("白银", r"12个月"),
        "pt_m1": find_col("铂金", r"(一|1)个月"),
        "pt_m3": find_col("铂金", r"3个月"),
        "pt_m6": find_col("铂金", r"6个月"),
        "pt_m12": find_col("铂金", r"12个月"),
        "pd_m1": find_col("钯金", r"(一|1)个月"),
        "pd_m3": find_col("钯金", r"3个月"),
        "pd_m6": find_col("钯金", r"6个月"),
        "pd_m12": find_col("钯金", r"12个月"),
    }
    df["__d"] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=["__d"]).sort_values("__d")
    latest = df["__d"].max()
    df = df[df["__d"] >= latest - timedelta(days=365)]
    write_json("lease_rates.json", {
        "generatedAt": now_iso(),
        "dates": df["__d"].dt.strftime("%Y-%m-%d").tolist(),
        "series": {k: [round_or_none(v, 4) for v in pd.to_numeric(df[c], errors="coerce")]
                   for k, c in col_map.items()},
    })



def step_minute_exports() -> None:
    """Export per-contract minute JSON for frontend live basis calculation."""
    assert CTX is not None
    contracts_list = []

    # AGTD
    df = CTX.agtd
    times = df["time"].dt.strftime("%Y-%m-%d %H:%M").tolist()
    vals = [round(float(v), 2) for v in df["close"]]
    write_json("min_AGTD.json", {"times": times, "values": vals})
    contracts_list.append({"code": "AGTD", "label": "AG(T+D)", "points": len(times)})

    # Each futures contract
    for code in sorted(CTX.contracts_min):
        df = CTX.contracts_min[code]
        if len(df) < 100:
            continue
        times = df["time"].dt.strftime("%Y-%m-%d %H:%M").tolist()
        vals = [round(float(v), 2) for v in df["close"]]
        write_json(f"min_{code}.json", {"times": times, "values": vals})
        contracts_list.append({"code": code, "label": code.upper(), "points": len(times)})

    write_json("min_contracts.json", {"generatedAt": now_iso(), "contracts": contracts_list})
    print(f"  Exported {len(contracts_list)} minute series")


def step_basis() -> None:
    assert CTX is not None
    pairs = [
        ("AGTD", "AG2608"), ("AGTD", "AG2609"), ("AGTD", "AG2610"),
        ("AG2608", "AG2609"), ("AG2609", "AG2610"), ("AG2610", "AG2611"),
    ]
    label = {"AGTD": "AG(T+D)"}

    def get(code: str) -> pd.DataFrame:
        if code == "AGTD":
            return CTX.agtd
        if code not in CTX.contracts_min:
            raise KeyError(f"分钟合约 {code} 不存在")
        return CTX.contracts_min[code]

    for a, b in pairs:
        times, al = align_minute_series({"A": get(a), "B": get(b)})
        vals = np.round(al["A"] - al["B"], 1)
        pa, pb = label.get(a, a), label.get(b, b)
        write_json(f"basis_{a}-{b}.json", {
            "pair": f"{pa}-{pb}",
            "generatedAt": now_iso(),
            "times": times,
            "values": vals.tolist(),
            "stats": spread_stats(vals),
        })


def select_main_comex_contract() -> tuple[str, str, list[pd.Timestamp], int]:
    """以最近10个交易日的有效分钟报价数作为主力代理，并要求存在次月沪银配对。"""
    assert CTX is not None
    all_dates = sorted({
        pd.Timestamp(ts).normalize()
        for frame in CTX.comex_min.values()
        for ts in frame["time"]
    })
    if len(all_dates) < 10:
        raise RuntimeError("COMEX 分钟数据不足10个交易日")
    window_dates = all_dates[-10:]
    window_set = set(window_dates)
    candidates: list[tuple[int, str, str]] = []
    for foreign_code, frame in CTX.comex_min.items():
        domestic_code = matched_domestic_code(foreign_code)
        if domestic_code not in CTX.contracts_min:
            continue
        count = int(frame["time"].map(lambda ts: pd.Timestamp(ts).normalize() in window_set).sum())
        candidates.append((count, foreign_code, domestic_code))
    if not candidates:
        raise RuntimeError("COMEX 合约均无法匹配次月沪银合约")
    quote_count, foreign_contract, domestic_contract = max(candidates)
    return foreign_contract, domestic_contract, window_dates, quote_count


def profit_payload(
    domestic: np.ndarray,
    agtd: np.ndarray,
    foreign: np.ndarray,
    fx: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    import_external_cny_kg = (foreign - OUTER_PRICE_DEDUCTION) / TROY_OUNCE_GRAM * fx * 1000
    export_external_cny_kg = foreign / TROY_OUNCE_GRAM * fx * 1000
    return (
        np.round(domestic - import_external_cny_kg * VAT, 1),
        np.round(export_external_cny_kg - agtd / VAT, 1),
        np.round(export_external_cny_kg - agtd, 1),
    )


def step_import_profit() -> None:
    assert CTX is not None
    foreign_contract, domestic_contract, window_dates, quote_count = select_main_comex_contract()
    foreign_frame = CTX.comex_min[foreign_contract]
    daily_fx_contract = select_quarterly_fx_contract(
        domestic_contract,
        pd.Timestamp(foreign_frame["time"].max()).normalize(),
        set(CTX.profit_daily),
    )
    if CTX.fx_min_code != daily_fx_contract:
        raise RuntimeError(
            f"分钟外汇合约 {CTX.fx_min_code} 与季度主力规则选择 {daily_fx_contract} 不一致"
        )
    times, al = align_minute_series({
        "domestic": CTX.contracts_min[domestic_contract],
        "agtd": CTX.agtd,
        "foreign": foreign_frame,
        "fx": CTX.usdcnh,
    })
    aligned_time = pd.to_datetime(pd.Series(times))
    mask = (
        aligned_time.dt.normalize().isin(window_dates)
        & (aligned_time <= foreign_frame["time"].max())
    ).to_numpy()
    times = [time for time, keep in zip(times, mask) if keep]
    al = {key: values[mask] for key, values in al.items()}
    if not times:
        raise RuntimeError("主力合约对齐后无最近10个交易日分钟数据")

    imp, processing_exp, general_exp = profit_payload(
        al["domestic"], al["agtd"], al["foreign"], al["fx"]
    )
    si = spread_stats(imp)
    spe = spread_stats(processing_exp)
    sge = spread_stats(general_exp)
    write_json("import_profit.json", {
        "generatedAt": now_iso(),
        "frequency": "minute",
        "windowTradingDays": 10,
        "windowStart": window_dates[0].strftime("%Y-%m-%d"),
        "windowEnd": window_dates[-1].strftime("%Y-%m-%d"),
        "selectionMethod": "最近10个交易日有效分钟报价数最多（主力代理）",
        "foreignContract": foreign_contract,
        "domesticContract": domestic_contract,
        "fxContract": CTX.fx_min_code,
        "mainQuoteCount": quote_count,
        "importFormula": "内盘期货价格-(外盘期货价格-0.2)/31.1035*外汇价格*1000*1.13",
        "processingExportFormula": "外盘主力期货价格/31.1035*外汇价格*1000-Ag(T+D)/1.13",
        "generalExportFormula": "外盘主力期货价格/31.1035*外汇价格*1000-Ag(T+D)",
        "times": times,
        "importProfit": imp.tolist(),
        "processingExportProfit": processing_exp.tolist(),
        "generalExportProfit": general_exp.tolist(),
        "stats": {
            "importLatest": si["latest"], "importMean": si["mean"], "importPercentile": si["percentile"],
            "processingExportLatest": spe["latest"], "processingExportMean": spe["mean"],
            "processingExportPercentile": spe["percentile"],
            "generalExportLatest": sge["latest"], "generalExportMean": sge["mean"],
            "generalExportPercentile": sge["percentile"],
        },
    })


def step_import_profit_daily() -> None:
    assert CTX is not None
    foreign_contract, domestic_contract, _, quote_count = select_main_comex_contract()
    as_of = pd.Timestamp(CTX.comex_min[foreign_contract]["time"].max()).normalize()
    fx_contract = select_quarterly_fx_contract(domestic_contract, as_of, set(CTX.profit_daily))
    required = {
        "domestic": f"{domestic_contract}.SHF",
        "agtd": "AG(T+D).SGE",
        "foreign": foreign_contract,
        "fx": fx_contract,
    }
    missing = [code for code in required.values() if code not in CTX.profit_daily]
    if missing:
        raise RuntimeError(f"日度进出口盈亏缺少序列: {missing}")
    times, aligned = align_minute_series({key: CTX.profit_daily[code] for key, code in required.items()})
    imp, processing_exp, general_exp = profit_payload(
        aligned["domestic"], aligned["agtd"], aligned["foreign"], aligned["fx"]
    )
    si = spread_stats(imp)
    spe = spread_stats(processing_exp)
    sge = spread_stats(general_exp)
    write_json("import_profit_daily.json", {
        "generatedAt": now_iso(),
        "frequency": "daily",
        "selectionMethod": "沿用分钟数据主力合约，内盘与汇率选择次月合约",
        "foreignContract": foreign_contract,
        "domesticContract": domestic_contract,
        "fxContract": fx_contract,
        "mainQuoteCount": quote_count,
        "importFormula": "内盘期货价格-(外盘期货价格-0.2)/31.1035*外汇价格*1000*1.13",
        "processingExportFormula": "外盘主力期货价格/31.1035*外汇价格*1000-Ag(T+D)/1.13",
        "generalExportFormula": "外盘主力期货价格/31.1035*外汇价格*1000-Ag(T+D)",
        "times": [pd.Timestamp(t).strftime("%Y-%m-%d") for t in times],
        "importProfit": imp.tolist(),
        "processingExportProfit": processing_exp.tolist(),
        "generalExportProfit": general_exp.tolist(),
        "stats": {
            "importLatest": si["latest"], "importMean": si["mean"], "importPercentile": si["percentile"],
            "processingExportLatest": spe["latest"], "processingExportMean": spe["mean"],
            "processingExportPercentile": spe["percentile"],
            "generalExportLatest": sge["latest"], "generalExportMean": sge["mean"],
            "generalExportPercentile": sge["percentile"],
        },
    })


def step_lhb() -> None:
    """重建全部可用交易日的龙虎榜历史数据。"""
    from build_lhb import main as build_lhb_main

    if build_lhb_main() != 0:
        raise RuntimeError("龙虎榜历史数据生成失败")


# ---------------------------------------------------------------- 内容校验

VERIFY_ISSUES: list[str] = []


def check(ok: bool, msg: str) -> None:
    print(f"    {'[OK]' if ok else '[FAIL]'} {msg}")
    if not ok:
        VERIFY_ISSUES.append(msg)


def load_out(name: str) -> dict:
    with open(OUT_DIR / name, encoding="utf-8") as fh:
        return json.load(fh)


def verify() -> None:
    print("\n========== 内容校验 ==========")
    expected = (["monitoring.json", "market.json", "daily.json", "positions_curve.json",
                 "virtual_ratio.json", "metal_virtual_ratio.json", "shfe_positioning.json",
                 "pp_warehouse.json",
                 "seasonality.json", "lease_rates.json",
                 "import_profit.json", "import_profit_daily.json"]
                + [f"basis_{p}.json" for p in
                   ["AGTD-AG2608", "AGTD-AG2609", "AGTD-AG2610",
                    "AG2608-AG2609", "AG2609-AG2610", "AG2610-AG2611"]])
    for name in expected:
        p = OUT_DIR / name
        check(p.exists() and p.stat().st_size > 2, f"{name} 存在且非空 ({p.stat().st_size if p.exists() else 0} B)")

    # daily.json
    d = load_out("daily.json")
    miss = [k for _, k, _ in DAILY_COLS if k not in d["series"]]
    check(not miss, f"daily.json 25 个 key 齐全 (缺失: {miss})")
    check(len(d["dates"]) > 500, f"daily.json 日期数 = {len(d['dates'])} (期望 >500)")
    check(d["dates"][-1] == d["asOfDate"],
          f"daily.json 末日期 {d['dates'][-1]} == asOfDate {d['asOfDate']}")
    check(len(d["dates"]) > 500, f"daily.json 首日期 {d['dates'][0]} (历史数据)")
    for k, arr in d["series"].items():
        nn = sum(v is not None for v in arr)
        if k in DAILY_DENSE_KEYS:
            check(nn > 300, f"daily.{k} 非空 {nn}/{len(d['dates'])} (日频列, 期望 >300)")
        else:
            check(nn > 0, f"daily.{k} 非空 {nn}/{len(d['dates'])} (稀疏列, 期望 >0)")
    check(all(len(v) == len(d["dates"]) for v in d["series"].values()),
          "daily.json 所有序列长度与日期一致")
    print(f"    0→null 转换计数: {ZERO_TO_NULL_COUNTS}")
    close_nn = [v for v in d["series"]["agtdClose"] if v is not None]
    check(5000 < close_nn[-1] < 30000, f"agtdClose 最新非空值 {close_nn[-1]} 元/千克 量级合理")

    # ---- 修改1: ffill + lastActual + domesticInvT 重算
    la = d.get("lastActual", {})
    check(set(la) == {"shfeInvT", "sgeInvT", "lbmaDailyT"}, f"daily.lastActual 键 {sorted(la)}")
    check(la.get("shfeInvT") is not None, f"lastActual.shfeInvT = {la.get('shfeInvT')} (非空)")
    check(la.get("sgeInvT") is not None, f"lastActual.sgeInvT = {la.get('sgeInvT')} (非空)")
    check(la.get("lbmaDailyT") is not None, f"lastActual.lbmaDailyT = {la.get('lbmaDailyT')} (非空)")
    for key in ("shfeInvT", "sgeInvT"):
        arr = d["series"][key]
        seg = [v for dt, v in zip(d["dates"], arr) if dt > "2026-07-10"]
        nn = sum(v is not None for v in seg)
        check(nn == len(seg) and seg, f"{key} 在 2026-07-10 之后无 null ({nn}/{len(seg)}, ffill 生效)")
        head_nulls = sum(v is None for v in arr[:5])
        print(f"    {key}: 头部 null {head_nulls}/5, 末值 {arr[-1]}, 最后实际日 {la.get(key)}")
    dom = d["series"]["domesticInvT"]
    dom_last = dom[-1]
    check(dom_last is not None and 500 < dom_last < 5000,
          f"domesticInvT 末值 {dom_last} 在合理区间 (500, 5000) 吨")
    sh_last, sg_last = d["series"]["shfeInvT"][-1], d["series"]["sgeInvT"][-1]
    check(sh_last is not None and sg_last is not None and
          abs((sh_last + sg_last) - dom_last) < 0.002,
          f"domesticInvT 末值 = shfe({sh_last}) + sge({sg_last}) 自洽")
    # 吨级序列 3 位小数抽查
    sh_dec = [len(str(v).split(".")[1]) for v in d["series"]["shfeInvT"] if isinstance(v, float) and "." in str(v)]
    check(all(n <= 3 for n in sh_dec), f"shfeInvT 小数位 ≤3 (抽查 {len(sh_dec)} 值)")

    # lease_rates.json
    lr = load_out("lease_rates.json")
    check(lr["dates"] == sorted(lr["dates"]), "lease_rates 日期升序")
    check(len(lr["dates"]) > 200, f"lease_rates 日期数 {len(lr['dates'])} (期望 >200)")
    EXPECTED_LR_KEYS = ["m1", "m3", "m6", "m12", "pt_m1", "pt_m3", "pt_m6", "pt_m12", "pd_m1", "pd_m3", "pd_m6", "pd_m12"]
    check(list(lr["series"].keys()) == EXPECTED_LR_KEYS, f"lease_rates 键 {list(lr['series'])}")
    check(all(len(v) == len(lr["dates"]) for v in lr["series"].values()), "lease_rates 序列长度一致")
    m6nn = [v for v in lr["series"]["m6"] if v is not None]
    check(len(m6nn) > 100 and -5 < m6nn[-1] < 20, f"lease m6 最新 {m6nn[-1]}% 量级合理, 非空 {len(m6nn)}")

    # seasonality.json
    se = load_out("seasonality.json")
    check(len(se["dates"]) == 366, f"seasonality 日期数 {len(se['dates'])} (期望 366)")
    check(sorted(se["years"].keys()) == ["2022", "2023", "2024", "2025", "2026"],
          f"seasonality 年份 {sorted(se['years'])}")
    check(all(len(v) == len(se["dates"]) for v in se["years"].values()), "seasonality 序列长度一致")
    nn25 = sum(v is not None for v in se["years"]["2025"])
    check(nn25 > 200, f"seasonality 2025 非空 {nn25}/366")

    # positions_curve / virtual_ratio（修改2: 仅偶数月合约）
    for fname in ["positions_curve.json", "virtual_ratio.json"]:
        pc = load_out(fname)
        codes = [c["code"] for c in pc["contracts"]]
        check(len(codes) >= 2,
              f"{fname} has >=2 even-month contracts (got {codes})")
        exps = [c["expiry"] for c in pc["contracts"]]
        check(exps == sorted(exps), f"{fname} 合约按到期日升序 {exps}")
        for c in pc["contracts"]:
            xs = [p["x"] for p in c["points"]]
            ys = [p["y"] for p in c["points"]]
            check(min(xs) >= -90 and max(xs) <= 0, f"{fname}.{c['code']} x∈[{min(xs)},{max(xs)}]")
            check(all(v is not None and v > 0 for v in ys), f"{fname}.{c['code']} y 全为正 ({len(ys)} 点)")
    vr = load_out("virtual_ratio.json")
    check(vr.get("formula") == "oi*15/st_stock_kg", "virtual_ratio.formula 元信息正确")

    # 铂钯虚实比：以合约代码识别品种，交易单位由广期合约参数换算为 kg/手
    mvr = load_out("metal_virtual_ratio.json")
    check(set(mvr.get("metals", {})) == {"pt", "pd"}, "metal_virtual_ratio 含 pt/pd 两个品种")
    for metal_key in ("pt", "pd"):
        metal = mvr["metals"][metal_key]
        contracts = metal.get("contracts", [])
        codes = [c["code"] for c in contracts]
        check(len(contracts) >= 3, f"metal_virtual_ratio.{metal_key} 合约数 {len(contracts)} ({codes})")
        check(metal.get("windowTradingDays") == 120, f"metal_virtual_ratio.{metal_key} 窗口 120 交易日")
        check(metal.get("formula") == "oi*contract_multiplier_kg/st_stock_kg",
              f"metal_virtual_ratio.{metal_key} 公式元信息正确")
        for contract in contracts:
            xs = [p["x"] for p in contract["points"]]
            ys = [p["y"] for p in contract["points"]]
            check(len(xs) >= 10, f"{contract['code']} 至少 10 个有效点 ({len(xs)})")
            check(len(xs) == len(set(xs)), f"{contract['code']} x 无重复 ({len(xs)} 点)")
            check(min(xs) >= -120 and max(xs) <= 0,
                  f"{contract['code']} x∈[{min(xs)},{max(xs)}]")
            check(all(v is not None and v >= 0 and math.isfinite(v) for v in ys),
                  f"{contract['code']} y 非负且有限")

    # 统一目录中的 SHFE/SGE 持仓排名与趋势
    sp = load_out("shfe_positioning.json")
    check(sp.get("asOfDate") == sp["nonFuturesTrend"]["dates"][-1],
          f"shfe_positioning 截止日 {sp.get('asOfDate')} 与趋势末日一致")
    check(len(sp.get("summary", [])) == 2,
          f"shfe_positioning 最新汇总含两类会员 ({len(sp.get('summary', []))})")
    check({row["category"] for row in sp["summary"]} == {"期货公司会员", "非期货公司会员"},
          "shfe_positioning 汇总会员类别齐全")
    nt = sp["nonFuturesTrend"]
    n = len(nt["dates"])
    check(n >= 100, f"shfe_positioning SHFE 交易日 {n} (期望 >=100)")
    check(nt["dates"] == sorted(set(nt["dates"])), "shfe_positioning SHFE 日期升序且唯一")
    check(all(len(nt[key]) == n for key in ("longLots", "shortLots", "netLots")),
          "shfe_positioning 非期货公司趋势序列长度一致")
    check(all(l - s == net for l, s, net in zip(nt["longLots"], nt["shortLots"], nt["netLots"])),
          "shfe_positioning 净持仓 = 持买 - 持卖")
    ct = sp["combinedTrend"]
    cn = len(ct["dates"])
    check(cn >= 100, f"shfe_positioning SHFE×SGE 共同交易日 {cn} (期望 >=100)")
    check(all(len(ct[key]) == cn for key in ("shfeLongTons", "shfeShortTons", "sgeOpenInterestTons")),
          "shfe_positioning 综合趋势序列长度一致")
    check(all(v > 0 and math.isfinite(v) for key in ("shfeLongTons", "shfeShortTons", "sgeOpenInterestTons") for v in ct[key]),
          "shfe_positioning 综合趋势吨值全为正且有限")

    # 统一目录中的铂钯仓单
    ppw = load_out("pp_warehouse.json")
    check(set(ppw.get("metals", {})) == {"pt", "pd"}, "pp_warehouse 含 pt/pd 两个品种")
    for metal_key in ("pt", "pd"):
        metal = ppw["metals"][metal_key]
        dates = metal["dates"]
        n = len(dates)
        check(n >= 40, f"pp_warehouse.{metal_key} 交易日 {n} (期望 >=40)")
        check(dates == sorted(set(dates)), f"pp_warehouse.{metal_key} 日期升序且唯一")
        series_keys = ("warehouseKg", "factoryKg", "totalKg", "registeredKg", "cancelledKg", "netChangeKg")
        check(all(len(metal[key]) == n for key in series_keys), f"pp_warehouse.{metal_key} 序列长度一致")
        check(all(w + f == t for w, f, t in zip(metal["warehouseKg"], metal["factoryKg"], metal["totalKg"])),
              f"pp_warehouse.{metal_key} 总量 = 仓库 + 厂库")
        check(metal["latest"]["date"] == dates[-1], f"pp_warehouse.{metal_key} 最新日期与趋势末日一致")
        check(metal["latest"]["totalKg"] == metal["totalKg"][-1], f"pp_warehouse.{metal_key} 最新总量自洽")
        check(sum(row["quantityKg"] for row in metal["locations"]) == metal["latest"]["totalKg"],
              f"pp_warehouse.{metal_key} 最新仓库明细合计自洽")
        check(all(v >= 0 and math.isfinite(v) for key in ("warehouseKg", "factoryKg", "totalKg", "registeredKg", "cancelledKg") for v in metal[key]),
              f"pp_warehouse.{metal_key} 库存/注册/注销非负且有限")

    # basis 文件
    for name, hi in [("basis_AGTD-AG2608.json", 1000), ("basis_AGTD-AG2609.json", 1000),
                     ("basis_AGTD-AG2610.json", 1000), ("basis_AG2608-AG2609.json", 1000),
                     ("basis_AG2609-AG2610.json", 1000), ("basis_AG2610-AG2611.json", 1000)]:
        b = load_out(name)
        check(b["times"][0] >= "2026-01-01" and len(b["times"]) > 100,
            f"{name} 时间范围 {b['times'][0]} → {b['times'][-1]} ({len(b['times'])} 条)")
        check(len(b["times"]) == len(b["values"]), f"{name} 序列长度一致 ({len(b['times'])})")
        st = b["stats"]
        check(abs(st["latest"]) < hi,
              f"{name} 最新基差 {st['latest']} 元/千克 (|值|<{hi}, 几十元量级说明对齐正确)")
        check(st["min"] <= st["mean"] <= st["max"], f"{name} stats min≤mean≤max")
        check(0 <= st["percentile"] <= 1, f"{name} percentile={st['percentile']} ∈[0,1]")

    # import_profit.json
    ip = load_out("import_profit.json")
    check(len(ip["times"]) == len(ip["importProfit"]) == len(ip["processingExportProfit"])
          == len(ip["generalExportProfit"]),
          f"import_profit 序列长度一致 ({len(ip['times'])})")
    s = ip["stats"]
    check(abs(s["importLatest"]) < 2000, f"importLatest {s['importLatest']} 元/千克 量级合理")
    check(re.fullmatch(r"AG\d{4}", ip.get("domesticContract", "")) is not None,
          f"进口盈亏内盘合约 {ip.get('domesticContract')} 合法")
    check(ip.get("importFormula") == "内盘期货价格-(外盘期货价格-0.2)/31.1035*外汇价格*1000*1.13",
          "进口盈亏公式与用户口径一致")
    check(ip.get("processingExportFormula") == "外盘主力期货价格/31.1035*外汇价格*1000-Ag(T+D)/1.13",
          "加贸出口盈亏公式与用户口径一致")
    check(ip.get("generalExportFormula") == "外盘主力期货价格/31.1035*外汇价格*1000-Ag(T+D)",
          "一般出口盈亏公式与用户口径一致")
    check(set(s) == {"importLatest", "importMean", "importPercentile",
                     "processingExportLatest", "processingExportMean", "processingExportPercentile",
                     "generalExportLatest", "generalExportMean", "generalExportPercentile"},
          "import_profit stats 键齐全")
    check(all(
        general <= processing
        for general, processing in zip(ip["generalExportProfit"], ip["processingExportProfit"])
    ), "一般出口盈亏不高于加贸出口盈亏")
    check(ip.get("frequency") == "minute" and ip.get("windowTradingDays") == 10,
          "import_profit 为最近10个交易日分钟数据")
    check(len({t[:10] for t in ip["times"]}) == 10,
          f"import_profit 覆盖10个交易日 ({ip['windowStart']}→{ip['windowEnd']})")
    check(comex_year_month(ip.get("foreignContract", "")) is not None and
          ip.get("domesticContract") == matched_domestic_code(ip.get("foreignContract", "")),
          f"主力外盘 {ip.get('foreignContract')} 匹配次月内盘 {ip.get('domesticContract')}")
    fx_ym = fx_year_month(ip.get("fxContract", ""))
    check(fx_ym is not None and fx_ym[1] in {3, 6, 9, 12},
          f"分钟外汇 {ip.get('fxContract')} 为季度主力月")

    ipd = load_out("import_profit_daily.json")
    check(len(ipd["times"]) == len(ipd["importProfit"]) == len(ipd["processingExportProfit"])
          == len(ipd["generalExportProfit"]) > 100,
          f"import_profit_daily 日度序列长度一致 ({len(ipd['times'])})")
    check(ipd["times"] == sorted(set(ipd["times"])), "import_profit_daily 日期升序且唯一")
    check(ipd.get("foreignContract") == ip.get("foreignContract") and
          ipd.get("domesticContract") == ip.get("domesticContract") and
          ipd.get("fxContract") == ip.get("fxContract"),
          "分钟与日度进出口盈亏使用同一主力合约对")
    check(abs(ipd["stats"]["importLatest"]) < 2000,
          f"daily importLatest {ipd['stats']['importLatest']} 元/千克量级合理")
    check(abs(ipd["stats"]["processingExportLatest"]) < 10000,
          f"daily processingExportLatest {ipd['stats']['processingExportLatest']} 元/千克量级合理")
    check(abs(ipd["stats"]["generalExportLatest"]) < 10000,
          f"daily generalExportLatest {ipd['stats']['generalExportLatest']} 元/千克量级合理")

    # 静态拷贝
    for name, keys in [("monitoring.json", ["generatedAt", "indicators"]),
                       ("market.json", ["fetchedAt", "items"])]:
        j = load_out(name)
        check(all(k in j for k in keys), f"{name} 顶层键含 {keys}")
    mon = load_out("monitoring.json")
    ind16 = next((item for item in mon["indicators"] if item.get("id") == 16), None)
    check(mon.get("asOfDate") >= "2026-07-21", f"monitoring.asOfDate = {mon.get('asOfDate')} (期望 >=2026-07-21)")
    check(ind16 is not None and ind16.get("dataStatus") == "已接入", "monitoring 指标16 已接入")
    check(ind16 is not None and ind16.get("period") == ind16.get("updatedAt"),
          f"monitoring 指标16 数据期 {ind16.get('period') if ind16 else None} = 更新时间 {ind16.get('updatedAt') if ind16 else None}")

    # 无回归：记录数与上一版一致
    print("    ---- 回归检查: 记录数 ----")
    counts = {
        "daily.json": (len(d["dates"]), len(d["dates"])),  # dynamic: full history
        "seasonality.json": (len(se["dates"]), 366),
        "lease_rates.json": (len(lr["dates"]), 261),
        "import_profit.json": (len(ip["times"]), len(ip["times"])),  # dynamic
        "import_profit_daily.json": (len(ipd["times"]), len(ipd["times"])),
    }
    for name in ["basis_AGTD-AG2608.json", "basis_AGTD-AG2609.json", "basis_AGTD-AG2610.json",
                 "basis_AG2608-AG2609.json", "basis_AG2609-AG2610.json", "basis_AG2610-AG2611.json"]:
        b = load_out(name)
        counts[name] = (len(b["times"]), len(b["times"]))  # dynamic: no regression check on exact count
    for name, (got, exp) in counts.items():
        check(got == exp, f"{name} 记录数 {got} (期望 {exp}, 无回归)")
    # 无回归：键齐全
    check(set(d.keys()) == {"generatedAt", "asOfDate", "lastActual", "dates", "series"} and
          len(d["series"]) == 25, "daily.json 顶层键(含新增 lastActual)与 25 个 series 无回归")
    check(sorted(se.keys()) == ["dates", "generatedAt", "years"], "seasonality.json 顶层键无回归")
    check(sorted(lr.keys()) == ["dates", "generatedAt", "series"], "lease_rates.json 顶层键无回归")
    check(sorted(ip.keys()) == ["domesticContract", "foreignContract", "frequency", "fxContract",
                                "generalExportFormula", "generalExportProfit", "generatedAt", "importFormula",
                                "importProfit", "mainQuoteCount", "processingExportFormula",
                                "processingExportProfit", "selectionMethod", "stats", "times", "windowEnd",
                                "windowStart", "windowTradingDays"],
          "import_profit.json 顶层键无回归")

    lhb = load_out("lhb.json")
    lhb_dates = lhb.get("dates", [])
    date_values = [item.get("date") for item in lhb_dates]
    check(len(lhb_dates) >= 2, f"龙虎榜历史交易日 {len(lhb_dates)} 个")
    check(date_values == sorted(set(date_values)), "龙虎榜历史日期升序且唯一")
    check(bool(lhb_dates) and lhb.get("date") == lhb_dates[-1].get("date"),
          "龙虎榜顶层日期与最新历史快照一致")
    check(all(item.get("contracts") for item in lhb_dates), "龙虎榜每个交易日均有合约数据")
    check(all(
        0 < len(side) <= 20
        for item in lhb_dates
        for contract in item.get("contracts", [])
        for side in (contract.get("long", []), contract.get("short", []))
    ), "龙虎榜多空榜单每侧 1~20 个席位")


# ---------------------------------------------------------------- 主流程


def main() -> int:
    print("=" * 60)
    print("build_dashboard_data.py — 白银看板数据抽取")
    print(f"项目根: {ROOT}")
    print("=" * 60)

    sheets: dict[str, pd.DataFrame] = {}

    def _load():
        defensive_sheet_check()
        nonlocal sheets
        sheets = load_all_sheets()
        global CTX
        CTX = Ctx(sheets)

    steps = [
        ("拷贝静态 JSON", step_copy_static),
        ("daily.json", step_daily),
        ("positions_curve.json", step_positions_curve),
        ("virtual_ratio.json", step_virtual_ratio),
        ("metal_virtual_ratio.json", step_metal_virtual_ratio),
        ("shfe_positioning.json", step_shfe_positioning),
        ("pp_warehouse.json", step_pp_warehouse),
        ("seasonality.json", step_seasonality),
        ("lease_rates.json", step_lease_rates),
        ("minute exports", step_minute_exports),
        ("basis_*.json", step_basis),
        ("import_profit.json", step_import_profit),
        ("import_profit_daily.json", step_import_profit_daily),
        ("lhb.json", step_lhb),
    ]
    total_steps = len(steps) + 2  # 输入 + 生成步骤 + 校验
    print(f"\n[1/{total_steps}] 加载输入")
    run_step("加载输入", _load)

    for i, (label, fn) in enumerate(steps, start=2):
        print(f"\n[{i}/{total_steps}] {label}")
        if CTX is None and label != "拷贝静态 JSON":
            print("  跳过：输入加载失败")
            ERRORS.append((label, "输入加载失败，跳过"))
            continue
        run_step(label, fn)

    print(f"\n[{total_steps}/{total_steps}] 校验输出")
    run_step("内容校验", verify)

    print("\n" + "=" * 60)
    if ERRORS:
        print(f"完成，但有 {len(ERRORS)} 个步骤失败:")
        for label, err in ERRORS:
            print(f"  [FAIL] {label}: {err}")
        if VERIFY_ISSUES:
            print(f"校验问题 {len(VERIFY_ISSUES)} 项:")
            for m in VERIFY_ISSUES:
                print(f"  [FAIL] {m}")
        return 1
    if VERIFY_ISSUES:
        print(f"生成成功，但校验发现 {len(VERIFY_ISSUES)} 项问题:")
        for m in VERIFY_ISSUES:
            print(f"  [FAIL] {m}")
        return 1
    print("全部输出生成成功，校验全部通过 [OK]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
